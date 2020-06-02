import os
import openpyxl
from openpyxl.styles import Font
import eyed3
from musicexcelinit import *

# Get the excel workbook and sheets
try:
    file = openpyxl.load_workbook('musiclibrary.xlsx')
except FileNotFoundError:
    print("Creating workbook...")
    file = openpyxl.Workbook()
    file.save('musiclibrary.xlsx')

thing = MusicExcelInit(file)
thing.getdirs()
thing.createsheets()
musicpath = thing.givedirec()
sheets = file.get_sheet_names()

# Create an instance of the music excel init class and get the current music library

# Create a list of all the folders in the main library
musicfolders = os.listdir(musicpath)

# The folder with the music and the sheets in the file have to be iterated through at the same time for this to work correctly
for sheet, folder in zip(sheets, musicfolders):
    # For every sheet, add these three column titles to the first three columns
    currentsheet = file.get_sheet_by_name(sheet)

    ft = Font
    titlecell = currentsheet['A1']
    titlecell.font = Font(bold=True)
    titlecell = "Title"

    artistcell = currentsheet['B1']
    artistcell.font = Font(bold=True)
    artistcell = "Artist"

    albumcell = currentsheet['C1']
    albumcell.font = Font(bold=True)
    albumcell = "Album"

    # Get the current folder in the Master Library folder so we can iterate through the files inside
    currentpath = musicpath + '/' + folder
    print(currentpath)
    currentfolder = os.listdir(currentpath)

    # Iterate through the files in the music folders and the three columns per sheet in the file at the same time
    for currentfile, rowA, rowB, rowC in zip(currentfolder, currentsheet.iter_rows(min_col=1, max_col=1, min_row=2, max_row=(len(currentfolder) + 1)), currentsheet.iter_rows(min_col=2, max_col=2, min_row=2, max_row=(len(currentfolder) + 1)), currentsheet.iter_rows(min_col=3, max_col=3, min_row=2, max_row=(len(currentfolder) + 1))):
        # Get the extension to determine if the current item is an mp3 file or a folder
        extension = currentfile[len(currentfile) - 4:len(currentfile)]

        if extension == ".mp3":
            # Get the path for the file
            tagpath = currentpath + "\\" + currentfile

            # Use eyed3 to create the strings for the title, artist and album tag
            tag = eyed3.load(tagpath)
            title = tag.tag.title
            # longesttitle = title

            artist = tag.tag.artist
            # longestartist = artist

            album = tag.tag.album
            # longestalbum = album

            # Add the title to row 1, the artist to row 2 and the album to row 3
            for cell in rowA:
                cell.value = title
                # Just set the column width after using the loop to get the longest stuff
            for cell in rowB:
                cell.value = artist
            for cell in rowC:
                cell.value = album
        else:
            # If the current item isn't an mp3 (it will be assumed that non-files are folders that are albums),
            # add the title of the folder to row 1, and the string "ALBUM" to row 2
            # Note that this will also assume that non-mp3 files like wma or flac are albums as well
            # Can't do shit about that since the tagging info isn't the same for those kinds of files and eyed3 can't be used
            for cell in rowA:
                cell.value = currentfile
            for cell in rowB:
                cell.value = "ALBUM"


file.save('musiclibrary.xlsx')
